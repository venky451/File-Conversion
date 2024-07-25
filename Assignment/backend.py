import os
import json
import openpyxl
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from tkinter import filedialog, messagebox

# Define the Excel file path
current_dir = os.getcwd()
excel_file = os.path.join(current_dir, "data.xlsx")

def save_data_to_excel(name, email, phone, branch_str):
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        wb.save(excel_file)
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append([name, email, phone, branch_str])
    wb.save(excel_file)

def read_data_from_excel():
    data = []
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append({
                'Name': row[0],
                'Email': row[1],
                'Phone': row[2],
                'Branch': row[3]
            })
    return data

def save_as_json(data, file_path):
    with open(file_path, 'w') as f:
        for entry in data:
            json.dump(entry, f)
            f.write('\n\n')  # Add a newline character after each entry and an extra newline for gap

def save_as_txt(data, file_path):
    with open(file_path, 'w') as f:
        for entry in data:
            f.write("\n".join(f"{key}: {value}" for key, value in entry.items()) + "\n\n")

def save_as_xml(data, file_path):
    root = ET.Element("Users")
    for entry in data:
        user = ET.SubElement(root, "User")
        for key, value in entry.items():
            ET.SubElement(user, key).text = value
        user.tail = '\n'  # Add a newline character after each user's data
    tree = ET.ElementTree(root)
    tree.write(file_path, encoding='utf-8', xml_declaration=True)

def download_data(file_format):
    data = read_data_from_excel()
    file_path = filedialog.asksaveasfilename(defaultextension=f".{file_format}", 
                                             filetypes=[(file_format.upper(), f"*.{file_format}")])
    if not file_path:
        return
    
    if file_format == 'json':
        save_as_json(data, file_path)
    elif file_format == 'txt':
        save_as_txt(data, file_path)
    elif file_format == 'xml':
        save_as_xml(data, file_path)
    elif file_format == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        for entry in data:
            ws.append([entry['Name'], entry['Email'], entry['Phone'], entry['Branch']])
        wb.save(file_path)
    
    messagebox.showinfo("Download", f"Downloaded as {file_format}")