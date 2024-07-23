import os
import json
import openpyxl
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from tkinter import filedialog, messagebox

# Path to the Excel file
excel_file = r"C:\Users\T.Venkateshwar Reddy\OneDrive\Desktop\data.xlsx"

def create_excel_file():
    """
    Creates an Excel file with a default sheet and headers if it doesn't already exist.
    """
    if not os.path.exists(excel_file):
        wb = Workbook()  # Create a new workbook
        ws = wb.active  # Get the active worksheet
        ws.title = "UserData"  # Set the title of the worksheet
        # Append the header row to the worksheet
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        wb.save(excel_file)  # Save the workbook to the file

def save_data_to_excel(name, email, phone, branch_str):
    """
    Appends user data to the existing Excel file.

    Parameters:
    - name (str): The user's name
    - email (str): The user's email address
    - phone (str): The user's phone number
    - branch_str (str): The user's branch as a comma-separated string
    """
    wb = openpyxl.load_workbook(excel_file)  # Load the existing workbook
    ws = wb.active  # Get the active worksheet
    # Append the new user data
    ws.append([name, email, phone, branch_str])
    wb.save(excel_file)  # Save the workbook

def save_as_json(data, file_path):
    """
    Saves data to a JSON file.

    Parameters:
    - data (list of dict): The data to save
    - file_path (str): The path to save the JSON file
    """
    with open(file_path, 'w') as f:
        json.dump(data, f)  # Write data to the JSON file

def save_as_txt(data, file_path):
    """
    Saves data to a text file, with each entry on a new line.

    Parameters:
    - data (list of dict): The data to save
    - file_path (str): The path to save the text file
    """
    with open(file_path, 'w') as f:
        for entry in data:
            f.write(f"{entry}\n")  # Write each entry to the text file

def save_as_xml(data, file_path):
    """
    Saves data to an XML file.

    Parameters:
    - data (list of dict): The data to save
    - file_path (str): The path to save the XML file
    """
    root = ET.Element("Users")  # Create the root XML element
    for entry in data:
        user = ET.SubElement(root, "User")  # Create a 'User' element for each entry
        for key, value in entry.items():
            ET.SubElement(user, key).text = value  # Add sub-elements with corresponding values
    tree = ET.ElementTree(root)  # Create an ElementTree object
    tree.write(file_path)  # Write the XML to the file

def download_data(file_format, data):
    """
    Prompts the user to select a file path and then saves the data in the selected format.

    Parameters:
    - file_format (str): The format to save the file in ('json', 'txt', 'xml', 'xlsx')
    - data (list of dict): The data to save
    """
    # Prompt the user to select a file path for saving
    file_path = filedialog.asksaveasfilename(defaultextension=f".{file_format}", 
                                             filetypes=[(file_format.upper(), f"*.{file_format}")])
    if not file_path:
        return  # Exit if no file path was selected
    
    # Save data in the selected format
    if file_format == 'json':
        save_as_json(data, file_path)
    elif file_format == 'txt':
        save_as_txt(data, file_path)
    elif file_format == 'xml':
        save_as_xml(data, file_path)
    elif file_format == 'xlsx':
        wb = Workbook()  # Create a new workbook
        ws = wb.active  # Get the active worksheet
        ws.title = "UserData"  # Set the title of the worksheet
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])  # Add the header row
        for entry in data:
            ws.append([entry['Name'], entry['Email'], entry['Phone'], entry['Branch']])  # Append each entry
        wb.save(file_path)  # Save the workbook to the file
    
    # Inform the user that the file has been downloaded
    messagebox.showinfo("Download", f"Downloaded as {file_format}")