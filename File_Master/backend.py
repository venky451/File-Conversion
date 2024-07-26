import os
import json
import openpyxl
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from tkinter import filedialog, messagebox

#importing the necessary modules


#Excel Sheet File path 
current_dir = os.getcwd()
excel_file = os.path.join(current_dir, "data.xlsx")


#Here Excel Sheet is like the database which stores all the entered details from the UI
def save_data_to_excel(name, email, phone, branch_str):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append([name, email, phone, branch_str])
    wb.save(excel_file)


#This Function is used to reaad data from the excel sheet 
def read_data_from_excel():
    data = []  #empty list
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append({                #data is appended into the list
                'Name': row[0],
                'Email': row[1],
                'Phone': row[2],
                'Branch': row[3]
            })
    return data

#the data is saved in the json format 

def save_as_json(data, file_path):
    with open(file_path, 'w') as f:
        for entry in data:
            json.dump(entry, f)
            f.write('\n\n')  # Add a newline character after each entry 

#the data is saved in the txt format 

def save_as_txt(data, file_path):
    with open(file_path, 'w') as f:
        for entry in data:
            f.write("\n".join(f"{key}: {value}" for key, value in entry.items()) + "\n\n")


#the data is saved in the xml format 

def save_as_xml(data, file_path):
    root = ET.Element("Users")
    for entry in data:
        user = ET.SubElement(root, "User")
        for key, value in entry.items():
            ET.SubElement(user, key).text = value
        user.tail = '\n\n'  # Add a newline character after each user's data
    tree = ET.ElementTree(root)
    tree.write(file_path, encoding='utf-8', xml_declaration=True)

#This Function is used to download the data in the respective Format

def download_data(file_format):
    data = read_data_from_excel()
    file_path = filedialog.asksaveasfilename(defaultextension=f".{file_format}", 
                                             filetypes=[(file_format.upper(), f"*.{file_format}")])    #user can give the name to the file 
    if not file_path:
        return
    
    if file_format == 'json':
        save_as_json(data, file_path)    #json format

    elif file_format == 'txt':
        save_as_txt(data, file_path)     #txt format

    elif file_format == 'xml':
        save_as_xml(data, file_path)      #xml format

    elif file_format == 'xlsx':           #xlsx format
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        for entry in data:
            ws.append([entry['Name'], entry['Email'], entry['Phone'], entry['Branch']])
        wb.save(file_path)
    
    messagebox.showinfo("Download", f"Downloaded as {file_format}")       #pop up message will be displayed after the file is downloaded